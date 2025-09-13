#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Report Manager module for Telegram bot reporting system.
Handles generation of Excel reports with subscription/unsubscription statistics,
retention analysis, and historical data export.
"""

import os
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from db.db import DatabaseManager
from utils.time_utils import (
    get_almaty_now, get_today_date_str, get_date_n_days_ago,
    format_datetime_for_report, get_week_start_date, get_month_start_date,
    format_time_period_ru
)
from utils.logging_conf import log_report_generation

logger = logging.getLogger(__name__)


class ReportManager:
    """
    Manages generation of various types of reports for the Telegram bot system.
    """
    
    def __init__(self, db_manager: DatabaseManager, reports_dir: str = "reports_output"):
        """
        Initialize Report Manager.
        
        Args:
            db_manager: Database manager instance
            reports_dir: Directory to store generated reports
        """
        self.db = db_manager
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
        
        logger.info(f"ReportManager initialized with reports directory: {self.reports_dir}")
    
    def generate_daily_report(self, target_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate daily subscription/unsubscription report.
        
        Args:
            target_date: Target date in YYYY-MM-DD format (default: yesterday)
            
        Returns:
            Dict with report data and file path
        """
        if target_date is None:
            target_date = get_date_n_days_ago(1)  # Yesterday by default
        
        log_report_generation("daily", target_date, "started")
        
        try:
            # Get daily statistics
            daily_stats = self.db.get_daily_stats(target_date)
            
            # Get detailed events for the day
            events = self.db.get_events_for_period(target_date, target_date)
            
            # Get retention data for recent periods
            retention_7d = self._get_retention_for_date(target_date, 7)
            retention_14d = self._get_retention_for_date(target_date, 14)
            retention_30d = self._get_retention_for_date(target_date, 30)
            
            report_data = {
                'date': target_date,
                'stats': daily_stats,
                'events': events,
                'retention': {
                    '7_days': retention_7d,
                    '14_days': retention_14d,
                    '30_days': retention_30d
                },
                'generated_at': get_almaty_now().isoformat()
            }
            
            # Generate Excel file
            excel_path = self._generate_daily_excel(report_data)
            report_data['excel_file'] = str(excel_path)
            
            log_report_generation("daily", target_date, "completed", 
                                  events_count=len(events), excel_file=excel_path)
            
            return report_data
            
        except Exception as e:
            log_report_generation("daily", target_date, "failed", error=str(e))
            logger.error(f"Failed to generate daily report: {e}")
            raise
    
    def generate_weekly_report(self, target_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate weekly subscription/unsubscription report.
        
        Args:
            target_date: Target date in YYYY-MM-DD format (default: last Monday)
            
        Returns:
            Dict with report data and file path
        """
        if target_date is None:
            # Get last Monday
            today = date.fromisoformat(get_today_date_str())
            days_since_monday = today.weekday()
            target_date = (today - timedelta(days=days_since_monday + 7)).isoformat()
        
        week_start = get_week_start_date(target_date)
        week_end = (date.fromisoformat(week_start) + timedelta(days=6)).isoformat()
        
        log_report_generation("weekly", f"{week_start} to {week_end}", "started")
        
        try:
            # Get weekly statistics
            weekly_stats = self.db.get_weekly_stats(week_start)
            
            # Get events for the week
            events = self.db.get_events_for_period(week_start, week_end)
            
            # Get daily breakdown
            daily_breakdown = []
            current_date = date.fromisoformat(week_start)
            for i in range(7):
                day_str = current_date.isoformat()
                day_stats = self.db.get_daily_stats(day_str)
                daily_breakdown.append({
                    'date': day_str,
                    'weekday': current_date.strftime('%A'),
                    'stats': day_stats
                })
                current_date += timedelta(days=1)
            
            report_data = {
                'week_start': week_start,
                'week_end': week_end,
                'stats': weekly_stats,
                'daily_breakdown': daily_breakdown,
                'events': events,
                'generated_at': get_almaty_now().isoformat()
            }
            
            # Generate Excel file
            excel_path = self._generate_weekly_excel(report_data)
            report_data['excel_file'] = str(excel_path)
            
            log_report_generation("weekly", f"{week_start} to {week_end}", "completed",
                                  events_count=len(events), excel_file=excel_path)
            
            return report_data
            
        except Exception as e:
            log_report_generation("weekly", f"{week_start} to {week_end}", "failed", error=str(e))
            logger.error(f"Failed to generate weekly report: {e}")
            raise
    
    def generate_monthly_report(self, target_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate monthly subscription/unsubscription report.
        
        Args:
            target_date: Target date in YYYY-MM-DD format (default: last month)
            
        Returns:
            Dict with report data and file path
        """
        if target_date is None:
            # Get first day of last month
            today = date.fromisoformat(get_today_date_str())
            if today.month == 1:
                last_month = today.replace(year=today.year - 1, month=12, day=1)
            else:
                last_month = today.replace(month=today.month - 1, day=1)
            target_date = last_month.isoformat()
        
        month_start = get_month_start_date(target_date)
        # Get last day of month
        next_month = date.fromisoformat(month_start) + timedelta(days=32)
        month_end = (next_month.replace(day=1) - timedelta(days=1)).isoformat()
        
        log_report_generation("monthly", f"{month_start} to {month_end}", "started")
        
        try:
            # Get monthly statistics
            monthly_stats = self.db.get_monthly_stats(month_start)
            
            # Get events for the month
            events = self.db.get_events_for_period(month_start, month_end)
            
            # Get weekly breakdown
            weekly_breakdown = self._get_weekly_breakdown_for_month(month_start, month_end)
            
            # Get retention analysis for the month
            retention_analysis = self._get_monthly_retention_analysis(month_start, month_end)
            
            report_data = {
                'month_start': month_start,
                'month_end': month_end,
                'stats': monthly_stats,
                'weekly_breakdown': weekly_breakdown,
                'retention_analysis': retention_analysis,
                'events': events,
                'generated_at': get_almaty_now().isoformat()
            }
            
            # Generate Excel file
            excel_path = self._generate_monthly_excel(report_data)
            report_data['excel_file'] = str(excel_path)
            
            log_report_generation("monthly", f"{month_start} to {month_end}", "completed",
                                  events_count=len(events), excel_file=excel_path)
            
            return report_data
            
        except Exception as e:
            log_report_generation("monthly", f"{month_start} to {month_end}", "failed", error=str(e))
            logger.error(f"Failed to generate monthly report: {e}")
            raise
    
    def generate_retention_report(self, retention_days: int = 7, target_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate detailed retention analysis report.
        
        Args:
            retention_days: Number of days for retention analysis
            target_date: Target date for analysis (default: today)
            
        Returns:
            Dict with retention analysis and file path
        """
        if target_date is None:
            target_date = get_today_date_str()
        
        log_report_generation("retention", f"{retention_days}d from {target_date}", "started")
        
        try:
            # Get retention statistics
            retention_stats = self.db.get_retention_stats(retention_days, target_date)
            
            # Get detailed retention data
            retention_details = self._get_detailed_retention_analysis(retention_days, target_date)
            
            # Get historical retention trends
            retention_trends = self._get_retention_trends(retention_days, target_date, 30)  # Last 30 days
            
            report_data = {
                'retention_days': retention_days,
                'target_date': target_date,
                'stats': retention_stats,
                'details': retention_details,
                'trends': retention_trends,
                'generated_at': get_almaty_now().isoformat()
            }
            
            # Generate Excel file
            excel_path = self._generate_retention_excel(report_data)
            report_data['excel_file'] = str(excel_path)
            
            log_report_generation("retention", f"{retention_days}d from {target_date}", "completed",
                                  excel_file=excel_path)
            
            return report_data
            
        except Exception as e:
            log_report_generation("retention", f"{retention_days}d from {target_date}", "failed", error=str(e))
            logger.error(f"Failed to generate retention report: {e}")
            raise
    
    def export_full_database(self) -> str:
        """
        Export complete database to Excel file.
        
        Returns:
            str: Path to generated Excel file
        """
        log_report_generation("full_export", "all_data", "started")
        
        try:
            # Get all data from database
            journal_data = self.db.get_journal_for_excel()
            retention_data = self.db.get_retention_checks_for_excel()
            user_stats = self.db.get_user_stats_summary()
            
            # Generate timestamp for filename
            timestamp = get_almaty_now().strftime("%Y%m%d_%H%M%S")
            filename = f"full_database_export_{timestamp}.xlsx"
            file_path = self.reports_dir / filename
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet safely after adding at least one new sheet
            active = wb.active
            sheets_added = False
            
            # Add journal data
            if journal_data:
                self._add_dataframe_to_workbook(wb, pd.DataFrame(journal_data), "Journal")
                sheets_added = True
            
            # Add retention data
            if retention_data:
                self._add_dataframe_to_workbook(wb, pd.DataFrame(retention_data), "Retention_Checks")
                sheets_added = True
            
            # Add user statistics
            if user_stats:
                self._add_dataframe_to_workbook(wb, pd.DataFrame(user_stats), "User_Statistics")
                sheets_added = True
            
            # If no data was added, create an empty sheet
            if not sheets_added:
                ws = wb.create_sheet("No_Data")
                ws['A1'] = "Нет данных для экспорта"
                ws['A2'] = f"Время создания: {get_almaty_now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Remove default sheet only if we added other sheets
            if active is not None and sheets_added:
                wb.remove(active)
            
            # Save workbook
            wb.save(file_path)
            
            log_report_generation("full_export", "all_data", "completed",
                                  journal_records=len(journal_data),
                                  retention_records=len(retention_data),
                                  excel_file=file_path)
            
            logger.info(f"Full database export completed: {file_path}")
            return str(file_path)
            
        except Exception as e:
            log_report_generation("full_export", "all_data", "failed", error=str(e))
            logger.error(f"Failed to export full database: {e}")
            raise
    
    def _get_retention_for_date(self, target_date: str, retention_days: int) -> Dict[str, Any]:
        """
        Get retention statistics for a specific date and period.
        
        Args:
            target_date: Date to check retention for
            retention_days: Retention period in days
            
        Returns:
            Dict with retention statistics
        """
        try:
            retention_stats = self.db.get_retention_stats(retention_days, target_date)
            return retention_stats
        except Exception as e:
            logger.warning(f"Failed to get retention data for {target_date} ({retention_days}d): {e}")
            return {'total_subscriptions': 0, 'retained': 0, 'not_retained': 0, 'retention_rate': 0.0}
    
    def _get_weekly_breakdown_for_month(self, month_start: str, month_end: str) -> List[Dict[str, Any]]:
        """
        Get weekly breakdown for a month period.
        
        Args:
            month_start: Start of month
            month_end: End of month
            
        Returns:
            List of weekly statistics
        """
        weekly_breakdown = []
        current_date = date.fromisoformat(month_start)
        end_date = date.fromisoformat(month_end)
        
        while current_date <= end_date:
            week_start = get_week_start_date(current_date.isoformat())
            week_end_date = date.fromisoformat(week_start) + timedelta(days=6)
            
            # Don't go beyond month end
            if week_end_date > end_date:
                week_end_date = end_date
            
            week_end = week_end_date.isoformat()
            
            try:
                week_stats = self.db.get_weekly_stats(week_start)
                weekly_breakdown.append({
                    'week_start': week_start,
                    'week_end': week_end,
                    'stats': week_stats
                })
            except Exception as e:
                logger.warning(f"Failed to get weekly stats for {week_start}: {e}")
            
            # Move to next week
            current_date = date.fromisoformat(week_start) + timedelta(days=7)
        
        return weekly_breakdown
    
    def _get_monthly_retention_analysis(self, month_start: str, month_end: str) -> Dict[str, Any]:
        """
        Get retention analysis for a month.
        
        Args:
            month_start: Start of month
            month_end: End of month
            
        Returns:
            Dict with retention analysis
        """
        try:
            # Analyze different retention periods
            retention_7d = self.db.get_retention_stats(7, month_end)
            retention_14d = self.db.get_retention_stats(14, month_end)
            retention_30d = self.db.get_retention_stats(30, month_end)
            
            return {
                '7_days': retention_7d,
                '14_days': retention_14d,
                '30_days': retention_30d
            }
        except Exception as e:
            logger.warning(f"Failed to get monthly retention analysis: {e}")
            return {'7_days': {}, '14_days': {}, '30_days': {}}
    
    def _get_detailed_retention_analysis(self, retention_days: int, target_date: str) -> List[Dict[str, Any]]:
        """
        Get detailed retention analysis with individual user data.
        
        Args:
            retention_days: Retention period
            target_date: Target date
            
        Returns:
            List of detailed retention records
        """
        try:
            # Get subscriptions that need retention check
            subscriptions = self.db.get_subscriptions_for_retention_check(retention_days, target_date)
            
            detailed_data = []
            for subscription in subscriptions:
                retention_result = self.db.check_user_retention(
                    subscription['id'], 
                    subscription['tg_user_id'], 
                    subscription['event_time']
                )
                
                detailed_data.append({
                    'journal_id': subscription['id'],
                    'tg_user_id': subscription['tg_user_id'],
                    'username': subscription.get('username'),
                    'subscription_date': subscription['event_time'],
                    'retention_result': retention_result,
                    'inviter_id': subscription.get('inviter_id')
                })
            
            return detailed_data
            
        except Exception as e:
            logger.warning(f"Failed to get detailed retention analysis: {e}")
            return []
    
    def _get_retention_trends(self, retention_days: int, target_date: str, days_back: int) -> List[Dict[str, Any]]:
        """
        Get retention trends over a period.
        
        Args:
            retention_days: Retention period
            target_date: End date
            days_back: Number of days to look back
            
        Returns:
            List of daily retention statistics
        """
        trends = []
        end_date = date.fromisoformat(target_date)
        
        for i in range(days_back):
            check_date = (end_date - timedelta(days=i)).isoformat()
            try:
                retention_stats = self.db.get_retention_stats(retention_days, check_date)
                trends.append({
                    'date': check_date,
                    'stats': retention_stats
                })
            except Exception as e:
                logger.warning(f"Failed to get retention trends for {check_date}: {e}")
        
        return list(reversed(trends))  # Oldest first
    
    def _generate_daily_excel(self, report_data: Dict[str, Any]) -> Path:
        """Generate Excel file for daily report."""
        from .excel_templates import DailyReportTemplate
        
        template = DailyReportTemplate()
        timestamp = get_almaty_now().strftime("%Y%m%d_%H%M%S")
        filename = f"daily_report_{report_data['date']}_{timestamp}.xlsx"
        file_path = self.reports_dir / filename
        
        template.generate(report_data, file_path)
        return file_path
    
    def _generate_weekly_excel(self, report_data: Dict[str, Any]) -> Path:
        """Generate Excel file for weekly report."""
        from .excel_templates import WeeklyReportTemplate
        
        template = WeeklyReportTemplate()
        timestamp = get_almaty_now().strftime("%Y%m%d_%H%M%S")
        filename = f"weekly_report_{report_data['week_start']}_to_{report_data['week_end']}_{timestamp}.xlsx"
        file_path = self.reports_dir / filename
        
        template.generate(report_data, file_path)
        return file_path
    
    def _generate_monthly_excel(self, report_data: Dict[str, Any]) -> Path:
        """Generate Excel file for monthly report."""
        from .excel_templates import MonthlyReportTemplate
        
        template = MonthlyReportTemplate()
        timestamp = get_almaty_now().strftime("%Y%m%d_%H%M%S")
        filename = f"monthly_report_{report_data['month_start']}_to_{report_data['month_end']}_{timestamp}.xlsx"
        file_path = self.reports_dir / filename
        
        template.generate(report_data, file_path)
        return file_path
    
    def _generate_retention_excel(self, report_data: Dict[str, Any]) -> Path:
        """Generate Excel file for retention report."""
        from .excel_templates import RetentionReportTemplate
        
        template = RetentionReportTemplate()
        timestamp = get_almaty_now().strftime("%Y%m%d_%H%M%S")
        filename = f"retention_report_{report_data['retention_days']}d_{report_data['target_date']}_{timestamp}.xlsx"
        file_path = self.reports_dir / filename
        
        template.generate(report_data, file_path)
        return file_path
    
    def _add_dataframe_to_workbook(self, wb: Workbook, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Add DataFrame to workbook as a new sheet.
        
        Args:
            wb: Workbook to add sheet to
            df: DataFrame to add
            sheet_name: Name of the sheet
        """
        ws = wb.create_sheet(title=sheet_name)
        
        # Add DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def get_report_summary(self, days_back: int = 30) -> Dict[str, Any]:
        """
        Get summary of reports generated in the last N days.
        
        Args:
            days_back: Number of days to look back
            
        Returns:
            Dict with report summary
        """
        try:
            start_date = get_date_n_days_ago(days_back)
            end_date = get_today_date_str()
            
            # Get basic statistics for the period
            period_stats = []
            current_date = date.fromisoformat(start_date)
            end_date_obj = date.fromisoformat(end_date)
            
            total_subscriptions = 0
            total_unsubscriptions = 0
            
            while current_date <= end_date_obj:
                day_str = current_date.isoformat()
                try:
                    daily_stats = self.db.get_daily_stats(day_str)
                    period_stats.append({
                        'date': day_str,
                        'stats': daily_stats
                    })
                    
                    total_subscriptions += daily_stats.get('total_subscriptions', 0)
                    total_unsubscriptions += daily_stats.get('total_unsubscriptions', 0)
                    
                except Exception as e:
                    logger.warning(f"Failed to get daily stats for {day_str}: {e}")
                
                current_date += timedelta(days=1)
            
            return {
                'period': f"{start_date} to {end_date}",
                'total_subscriptions': total_subscriptions,
                'total_unsubscriptions': total_unsubscriptions,
                'net_growth': total_subscriptions - total_unsubscriptions,
                'daily_stats': period_stats,
                'generated_at': get_almaty_now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Failed to get report summary: {e}")
            return {'error': str(e)}
    
    def cleanup_old_reports(self, days_to_keep: int = 30) -> int:
        """
        Clean up old report files.
        
        Args:
            days_to_keep: Number of days to keep reports
            
        Returns:
            int: Number of files deleted
        """
        try:
            cutoff_time = datetime.now() - timedelta(days=days_to_keep)
            deleted_count = 0
            
            for file_path in self.reports_dir.glob("*.xlsx"):
                if file_path.stat().st_mtime < cutoff_time.timestamp():
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        logger.info(f"Deleted old report: {file_path}")
                    except Exception as e:
                        logger.warning(f"Failed to delete {file_path}: {e}")
            
            logger.info(f"Cleanup completed: {deleted_count} files deleted")
            return deleted_count
            
        except Exception as e:
            logger.error(f"Failed to cleanup old reports: {e}")
            return 0