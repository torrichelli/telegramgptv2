#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified Report Manager that integrates the new Excel template with existing system.
Handles both the unified Excel file and maintains compatibility with current workflows.
"""

import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path

from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

from db.db import DatabaseManager
from reports.unified_excel_template import UnifiedExcelTemplate
from utils.time_utils import get_almaty_now, format_datetime_for_report
from utils.logging_conf import get_logger

logger = get_logger(__name__)


class UnifiedReportManager:
    """
    Unified Report Manager that manages the subscribers_report.xlsx file
    according to TZ specification and provides compatibility with existing system.
    """
    
    def __init__(self, db_manager: DatabaseManager, reports_dir: str = "reports_output"):
        """Initialize unified report manager."""
        self.db = db_manager
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
        
        # Initialize unified Excel template
        self.excel_template = UnifiedExcelTemplate(db_manager, reports_dir)
        
        logger.info(f"UnifiedReportManager initialized with reports directory: {self.reports_dir}")
    
    def add_event_to_history(self, event_data: Dict[str, Any]) -> None:
        """Add event to История sheet and update Статистика."""
        try:
            # Add to История sheet
            self.excel_template.add_history_event(event_data)
            
            # Update Статистика sheet
            self.excel_template.update_statistics_sheet()
            
            logger.info(f"Added event to unified Excel: {event_data.get('event_type')} for user {event_data.get('tg_user_id')}")
            
        except Exception as e:
            logger.error(f"Error adding event to unified Excel: {e}")
    
    def generate_daily_report(self, target_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate daily report according to TZ specification.
        
        Args:
            target_date: Target date in YYYY-MM-DD format (default: yesterday)
            
        Returns:
            Dict with report data and file path
        """
        if target_date is None:
            yesterday = get_almaty_now() - timedelta(days=1)
            target_date = yesterday.strftime("%Y-%m-%d")
        
        try:
            target_date_obj = datetime.strptime(target_date, "%Y-%m-%d").date()
            
            # Update Статистика sheet first
            self.excel_template.update_statistics_sheet()
            
            # Create daily report sheet
            report_result = self.excel_template.create_daily_report_sheet(target_date_obj)
            
            # Generate daily message
            daily_message = self.excel_template.get_daily_report_message(target_date_obj)
            
            return {
                'success': True,
                'target_date': target_date,
                'message': daily_message,
                'excel_file': self.excel_template.get_excel_file_path(),
                'sheet_created': not report_result.get('sheet_exists', False),
                'stats': report_result.get('stats', {})
            }
            
        except Exception as e:
            logger.error(f"Error generating daily report: {e}")
            return {
                'success': False,
                'error': str(e),
                'target_date': target_date
            }
    
    def get_daily_message_with_button(self, target_date: Optional[str] = None) -> tuple[str, InlineKeyboardMarkup]:
        """
        Get daily report message with download button according to TZ specification.
        
        Returns:
            Tuple of (message_text, keyboard_markup)
        """
        report_data = self.generate_daily_report(target_date)
        
        if not report_data['success']:
            error_message = f"❌ Ошибка генерации отчёта: {report_data.get('error', 'Неизвестная ошибка')}"
            return error_message, InlineKeyboardMarkup(inline_keyboard=[])
        
        message = report_data['message']
        
        # Create download button according to TZ
        download_button = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📤 Скачать Excel-файл", callback_data="download:unified_excel")]
        ])
        
        return message, download_button
    
    def get_excel_file_path(self) -> str:
        """Get path to the unified subscribers_report.xlsx file."""
        return self.excel_template.get_excel_file_path()
    
    def get_stats_summary(self) -> Dict[str, Any]:
        """Get current statistics summary from Статистика sheet."""
        try:
            stats_data = self.excel_template._get_current_statistics()
            
            total_inviters = len(stats_data)
            total_invited = sum(stat['total_invited'] for stat in stats_data)
            total_active = sum(stat['currently_subscribed'] for stat in stats_data)
            
            return {
                'total_inviters': total_inviters,
                'total_invited': total_invited,
                'total_active': total_active,
                'overall_retention': round((total_active / total_invited * 100) if total_invited > 0 else 0),
                'inviters_data': stats_data
            }
            
        except Exception as e:
            logger.error(f"Error getting stats summary: {e}")
            return {
                'total_inviters': 0,
                'total_invited': 0,
                'total_active': 0,
                'overall_retention': 0,
                'inviters_data': []
            }
    
    def export_excel_file(self) -> str:
        """
        Export current unified Excel file.
        Updates statistics and returns file path.
        """
        try:
            # Update statistics before export
            self.excel_template.update_statistics_sheet()
            
            file_path = self.excel_template.get_excel_file_path()
            logger.info(f"Excel file ready for export: {file_path}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error exporting Excel file: {e}")
            raise
    
    def handle_subscription_event(self, user_data: Dict[str, Any]) -> None:
        """Handle subscription event - add to История and update Статистика."""
        event_data = {
            'event_time': get_almaty_now().isoformat(),
            'event_type': 'subscription',
            'tg_user_id': user_data.get('tg_user_id'),
            'username': user_data.get('username', ''),
            'user_name': user_data.get('name', ''),
            'inviter_name': user_data.get('inviter_name', ''),
            'status': 'subscribed'
        }
        
        self.add_event_to_history(event_data)
    
    def handle_unsubscription_event(self, user_data: Dict[str, Any]) -> None:
        """Handle unsubscription event - add to История and update Статистика."""
        event_data = {
            'event_time': get_almaty_now().isoformat(),
            'event_type': 'unsubscription', 
            'tg_user_id': user_data.get('tg_user_id'),
            'username': user_data.get('username', ''),
            'user_name': user_data.get('name', ''),
            'inviter_name': user_data.get('inviter_name', ''),
            'status': 'unsubscribed'
        }
        
        self.add_event_to_history(event_data)
    
    def get_daily_sheet_exists(self, target_date: Optional[date] = None) -> bool:
        """Check if daily sheet already exists for given date."""
        if target_date is None:
            target_date = (get_almaty_now() - timedelta(days=1)).date()
        
        try:
            from openpyxl import load_workbook
            
            excel_file = self.excel_template.get_excel_file_path()
            if not Path(excel_file).exists():
                return False
            
            wb = load_workbook(excel_file)
            sheet_name = target_date.strftime("%d-%m-%Y")
            
            return sheet_name in wb.sheetnames
            
        except Exception as e:
            logger.error(f"Error checking daily sheet existence: {e}")
            return False