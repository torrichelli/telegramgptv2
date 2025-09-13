#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Subscribers Database Manager - Единый Excel файл subscribers_database.xlsx по ТЗ.
Реализует работу с одним файлом, который накапливает историю и обновляет статистику.
"""

import os
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from db.db import DatabaseManager
from utils.time_utils import get_almaty_now, format_datetime_for_report
from utils.logging_conf import get_logger

logger = get_logger(__name__)


class SubscribersDatabaseManager:
    """
    Менеджер единого Excel файла subscribers_database.xlsx согласно ТЗ.
    Обеспечивает:
    1. Лист История - накопительно дописывает новые события
    2. Лист Статистика - перезаписывает актуальные данные
    3. Листы YYYY-MM-DD - создает новые листы для ежедневных отчетов
    """
    
    def __init__(self, db_manager: DatabaseManager, reports_dir: str = "reports_output"):
        """Initialize subscribers database manager."""
        self.db = db_manager
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
        self.excel_file = self.reports_dir / "subscribers_database.xlsx"
        
        # Инициализировать файл если его нет
        self._ensure_file_exists()
        
        logger.info(f"SubscribersDatabaseManager initialized: {self.excel_file}")
    
    def _ensure_file_exists(self) -> None:
        """Создать Excel файл с листами История и Статистика если его нет."""
        if not self.excel_file.exists():
            logger.info("Creating new subscribers_database.xlsx with base sheets")
            
            wb = Workbook()
            
            # Удаляем дефолтный лист
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Создаем листы согласно ТЗ
            self._create_history_sheet(wb)
            self._create_statistics_sheet(wb)
            
            wb.save(self.excel_file)
            logger.info("✅ Created subscribers_database.xlsx with История and Статистика sheets")
    
    def _create_history_sheet(self, wb: Workbook) -> None:
        """Создать лист История с правильными заголовками согласно ТЗ."""
        ws = wb.create_sheet("История")
        
        # Заголовки согласно ТЗ
        headers = [
            "Дата/время подписки", "Дата/время выхода", "Никнейм", 
            "Username", "Кто пригласил", "Статус"
        ]
        
        # Добавляем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        # Автоподбор ширины колонок
        self._auto_adjust_columns(ws)
    
    def _create_statistics_sheet(self, wb: Workbook) -> None:
        """Создать лист Статистика с правильными заголовками согласно ТЗ."""
        ws = wb.create_sheet("Статистика")
        
        # Заголовки согласно ТЗ
        headers = [
            "Пригласивший", "Всего приглашено", "Подписаны сейчас", "Отписались"
        ]
        
        # Добавляем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        # Автоподбор ширины колонок
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Автоподбор ширины колонок по содержимому."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_history_event(self, event_data: Dict[str, Any]) -> None:
        """
        Добавить новое событие в лист История (дописать в конец, не трогать старые данные).
        
        Args:
            event_data: Данные события с полями:
                - event_time: время события
                - event_type: тип события (subscribe/unsubscribe)
                - tg_user_id: ID пользователя
                - username: @username пользователя
                - user_name: имя пользователя (никнейм)
                - inviter_name: кто пригласил
                - status: текущий статус
        """
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["История"]
            
            # Найти следующую пустую строку
            next_row = ws.max_row + 1
            
            # Обработка времени события
            event_time = event_data.get('event_time')
            if isinstance(event_time, str):
                try:
                    event_time = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
                except:
                    event_time = get_almaty_now()
            
            # Получаем user_id для всех веток
            user_id = str(event_data.get('tg_user_id', ''))
            
            # Форматирование согласно ТЗ
            if event_data.get('event_type') in ['subscribe', 'subscription', 'join']:
                # Подписка - заполняем дату подписки, дату выхода оставляем пустой
                subscription_time = event_time.strftime("%d.%m.%Y %H:%M")
                exit_time = ""
                status = "подписан"
            else:
                # Отписка - ищем существующую запись для обновления
                subscription_time = ""
                exit_time = event_time.strftime("%d.%m.%Y %H:%M")
                status = "вышел"
                
                # Исправленная логика поиска: колонка 4 = Username (@username)
                username_to_find = event_data.get('username', '')
                if username_to_find and not username_to_find.startswith('@'):
                    username_to_find = f'@{username_to_find}'
                
                # Попытаемся найти существующую строку для обновления даты выхода
                for row_num in range(2, ws.max_row + 1):
                    existing_username = str(ws.cell(row=row_num, column=4).value or '')  # Username колонка
                    
                    if username_to_find and existing_username == username_to_find:
                        # Обновляем существующую запись
                        ws.cell(row=row_num, column=2, value=exit_time)  # Дата выхода
                        ws.cell(row=row_num, column=6, value=status)  # Статус
                        wb.save(self.excel_file)
                        logger.info(f"Updated exit time for user {user_id} (@{username_to_find})")
                        return
            
            # Обработка username (обеспечиваем @ префикс)
            username = event_data.get('username', '')
            if username and not username.startswith('@'):
                username = f'@{username}'
            
            # Обработка имени пригласителя (русификация Unknown)
            inviter_name = event_data.get('inviter_name', '')
            if inviter_name == 'Unknown' or not inviter_name:
                inviter_name = 'Не указан'
            
            # Данные строки согласно ТЗ формату
            row_data = [
                subscription_time,  # Дата/время подписки
                exit_time,          # Дата/время выхода
                event_data.get('user_name', event_data.get('name', '')),  # Никнейм
                username,           # Username
                inviter_name,       # Кто пригласил
                status              # Статус
            ]
            
            # Добавляем данные в строку
            for col, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col, value=value)
            
            wb.save(self.excel_file)
            logger.info(f"Added history event: {event_data.get('event_type')} for user {user_id}")
            
        except Exception as e:
            logger.error(f"Error adding history event: {e}")
            raise
    
    def update_statistics_sheet(self) -> None:
        """
        Обновить лист Статистика - очистить и перезаписать актуальные данные.
        Подсчеты делаются на момент экспорта согласно ТЗ.
        """
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Статистика"]
            
            # Очищаем существующие данные (оставляем заголовки)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            
            # Получаем актуальную статистику из базы
            stats_data = self._get_current_statistics()
            
            # Добавляем данные согласно ТЗ формату
            for row_idx, inviter_stats in enumerate(stats_data, 2):
                # Нормализация имени пригласителя
                inviter_name = inviter_stats.get('inviter_name', 'Не указан')
                if inviter_name == 'Unknown' or not inviter_name:
                    inviter_name = 'Не указан'
                
                ws.cell(row=row_idx, column=1, value=inviter_name)  # Пригласивший
                ws.cell(row=row_idx, column=2, value=inviter_stats.get('total_invited', 0))  # Всего приглашено
                ws.cell(row=row_idx, column=3, value=inviter_stats.get('currently_subscribed', 0))  # Подписаны сейчас
                
                # Рассчитываем отписавшихся
                total_invited = inviter_stats.get('total_invited', 0)
                currently_subscribed = inviter_stats.get('currently_subscribed', 0)
                unsubscribed = max(0, total_invited - currently_subscribed)
                ws.cell(row=row_idx, column=4, value=unsubscribed)  # Отписались
            
            # Автоподбор колонок
            self._auto_adjust_columns(ws)
            
            wb.save(self.excel_file)
            logger.info("Updated Статистика sheet with current data")
            
        except Exception as e:
            logger.error(f"Error updating statistics sheet: {e}")
            raise
    
    def _get_current_statistics(self) -> List[Dict[str, Any]]:
        """Получить актуальную статистику по пригласителям из базы данных."""
        try:
            return self.db.get_statistics_data()
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return []
    
    def create_daily_report_sheet(self, target_date: date) -> Dict[str, Any]:
        """
        Создать новый лист с отчетом за день в формате YYYY-MM-DD в subscribers_database.xlsx.
        Если лист уже существует - пропускаем согласно ТЗ.
        
        Args:
            target_date: Дата для отчета
            
        Returns:
            Dict с результатом операции
        """
        try:
            sheet_name = target_date.strftime("%Y-%m-%d")
            
            wb = load_workbook(self.excel_file)
            
            # Проверяем есть ли уже такой лист
            if sheet_name in wb.sheetnames:
                logger.info(f"Daily sheet {sheet_name} already exists in subscribers_database.xlsx - skipping")
                return {'sheet_exists': True, 'sheet_name': sheet_name, 'file': str(self.excel_file)}
            
            # Создаем новый лист в subscribers_database.xlsx
            ws = wb.create_sheet(sheet_name)
            
            # Получаем данные за день
            daily_stats = self._get_daily_statistics(target_date)
            
            # Формируем отчет согласно ТЗ
            current_row = 1
            
            # Заголовок
            title_cell = ws.cell(row=current_row, column=1, value=f"Ежедневный отчет за {target_date.strftime('%d.%m.%Y')}")
            title_cell.font = Font(size=14, bold=True)
            current_row += 2
            
            # Текстовый отчет согласно ТЗ
            report_content = f"""Сводка по удержанию:
- Новых пользователей: {daily_stats['new_users']}
- Остались активными: {daily_stats['retained_users']} ({daily_stats['retention_rate']}%)
- Вышли в тот же день: {daily_stats['left_same_day']}

Топ-3 пригласителей:
{daily_stats['top_inviters_text']}

Динамика (vs вчера):
{daily_stats['dynamics_text']}"""
            
            # Добавляем текст отчета построчно для лучшего форматирования
            for line in report_content.split('\n'):
                if line.strip():
                    ws.cell(row=current_row, column=1, value=line)
                    current_row += 1
                else:
                    current_row += 1  # Пустая строка
            
            # Автоподбор колонок
            self._auto_adjust_columns(ws)
            
            wb.save(self.excel_file)
            logger.info(f"✅ Created daily report sheet {sheet_name} in subscribers_database.xlsx")
            
            return {
                'sheet_exists': False,
                'sheet_name': sheet_name,
                'stats': daily_stats,
                'file': str(self.excel_file)
            }
            
        except Exception as e:
            logger.error(f"Error creating daily report sheet: {e}")
            return {'error': str(e)}
    
    def _get_daily_statistics(self, target_date: date) -> Dict[str, Any]:
        """Получить статистику за конкретный день."""
        try:
            date_str = target_date.strftime("%Y-%m-%d")
            daily_stats = self.db.get_daily_stats(date_str)
            
            # Базовые показатели
            new_users = daily_stats.get('total_subscriptions', 0)
            left_same_day = daily_stats.get('same_day_unsubscriptions', 0)
            
            # Удержание через 3 дня
            retention_data = self.db.get_retention_for_date(date_str, 3)
            retained_users = retention_data.get('retained', 0)
            retention_rate = round(retention_data.get('retention_rate', 0))
            
            # Топ-3 пригласителей
            top_inviters = self._get_top_inviters_for_date(target_date)
            top_inviters_text = self._format_top_inviters(top_inviters)
            
            # Динамика (сравнение с вчера)
            yesterday = target_date - timedelta(days=1)
            yesterday_stats = self.db.get_daily_stats(yesterday.strftime("%Y-%m-%d"))
            dynamics_text = self._format_dynamics(daily_stats, yesterday_stats)
            
            return {
                'new_users': new_users,
                'retained_users': retained_users,
                'retention_rate': retention_rate,
                'left_same_day': left_same_day,
                'top_inviters_text': top_inviters_text,
                'dynamics_text': dynamics_text
            }
            
        except Exception as e:
            logger.error(f"Error getting daily statistics: {e}")
            return {
                'new_users': 0, 'retained_users': 0, 'retention_rate': 0,
                'left_same_day': 0, 'top_inviters_text': 'Нет данных',
                'dynamics_text': 'Нет данных'
            }
    
    def _get_top_inviters_for_date(self, target_date: date) -> List[Dict[str, Any]]:
        """Получить топ-3 пригласителей за день."""
        try:
            date_str = target_date.strftime("%Y-%m-%d")
            return self.db.get_top_inviters_for_date(date_str, limit=3)
        except:
            return []
    
    def _format_top_inviters(self, top_inviters: List[Dict[str, Any]]) -> str:
        """Форматировать топ пригласителей."""
        if not top_inviters:
            return "Нет данных"
        
        lines = []
        for i, inviter in enumerate(top_inviters, 1):
            name = inviter.get('inviter_name', 'Не указан')
            if name == 'Unknown':
                name = 'Не указан'
            invited = inviter.get('invited_count', 0)
            retained = inviter.get('retained_count', 0)
            retention_pct = round((retained / invited * 100) if invited > 0 else 0)
            
            lines.append(f"{i}. {name} — {invited} приглашено, {retained} удержано ({retention_pct}%)")
        
        return "\n".join(lines)
    
    def _format_dynamics(self, today_stats: Dict[str, Any], yesterday_stats: Dict[str, Any]) -> str:
        """Форматировать динамику сравнительно с вчера."""
        today_new = today_stats.get('total_subscriptions', 0)
        yesterday_new = yesterday_stats.get('total_subscriptions', 0)
        
        change = today_new - yesterday_new
        change_text = f"+{change}" if change > 0 else str(change)
        
        return f"Вчера: {yesterday_new} новых, Сегодня: {today_new} новых ({change_text})"
    
    def export_database(self) -> str:
        """
        Экспортировать единую базу данных согласно ТЗ.
        Обновляет все данные и возвращает путь к файлу.
        
        Returns:
            str: Путь к файлу subscribers_database.xlsx
        """
        try:
            # Сначала обновляем статистику
            self.update_statistics_sheet()
            
            # Создаем лист за сегодня если его еще нет
            today = get_almaty_now().date()
            self.create_daily_report_sheet(today)
            
            logger.info(f"Database export completed: {self.excel_file}")
            return str(self.excel_file)
            
        except Exception as e:
            logger.error(f"Error exporting database: {e}")
            raise
    
    def get_file_path(self) -> str:
        """Получить путь к файлу subscribers_database.xlsx."""
        return str(self.excel_file)