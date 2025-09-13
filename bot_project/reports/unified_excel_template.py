#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified Excel template for subscribers_report.xlsx according to TZ specification.
Creates and maintains a single Excel file with История, Статистика, and daily sheets.
"""

import os
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from db.db import DatabaseManager
from utils.time_utils import get_almaty_now, format_datetime_for_report
from utils.logging_conf import get_logger

logger = get_logger(__name__)


class UnifiedExcelTemplate:
    """
    Unified Excel template that maintains subscribers_report.xlsx 
    with История, Статистика, and daily report sheets.
    """
    
    def __init__(self, db_manager: DatabaseManager, reports_dir: str = "reports_output"):
        """Initialize unified Excel template."""
        self.db = db_manager
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
        self.excel_file = self.reports_dir / "subscribers_report.xlsx"
        
        # Initialize Excel file if it doesn't exist
        self._ensure_excel_file_exists()
        
        logger.info(f"UnifiedExcelTemplate initialized: {self.excel_file}")
    
    def _ensure_excel_file_exists(self) -> None:
        """Create Excel file with История and Статистика sheets if it doesn't exist."""
        if not self.excel_file.exists():
            logger.info("Creating new subscribers_report.xlsx with base sheets")
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create История sheet
            self._create_history_sheet(wb)
            
            # Create Статистика sheet  
            self._create_statistics_sheet(wb)
            
            wb.save(self.excel_file)
            logger.info("✅ Created subscribers_report.xlsx with История and Статистика sheets")
    
    def _create_history_sheet(self, wb: Workbook) -> None:
        """Create История sheet with proper headers."""
        ws = wb.create_sheet("История")
        
        # Headers according to TZ
        headers = [
            "Дата/время", "Действие", "User ID", "Username", 
            "Имя", "Пригласивший", "Статус"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_statistics_sheet(self, wb: Workbook) -> None:
        """Create Статистика sheet with proper headers."""
        ws = wb.create_sheet("Статистика")
        
        # Headers according to TZ
        headers = [
            "Пригласивший", "Всего приглашено", "Подписаны сейчас", 
            "Отписались", "% удержания"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_history_event(self, event_data: Dict[str, Any]) -> None:
        """Add new event to История sheet."""
        wb = load_workbook(self.excel_file)
        ws = wb["История"]
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Format datetime according to TZ (ДД.ММ.ГГГГ HH:MM)
        event_time = event_data.get('event_time')
        if isinstance(event_time, str):
            event_time = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
        
        formatted_time = event_time.strftime("%d.%m.%Y %H:%M")
        
        # Map event type to Russian according to TZ
        action_map = {
            'subscribe': 'подписка',
            'unsubscribe': 'отписка',
            'subscription': 'подписка',
            'unsubscription': 'отписка',
            'join': 'подписка',
            'leave': 'отписка'
        }
        action = action_map.get(event_data.get('event_type', ''), event_data.get('event_type', ''))
        
        # Map status to Russian according to TZ
        status_map = {
            'subscribed': 'подписан',
            'unsubscribed': 'вышел',
            'left': 'вышел',
            'active': 'подписан',
            'inactive': 'вышел'
        }
        status = status_map.get(event_data.get('status', ''), event_data.get('status', ''))
        
        # Format username (ensure @ prefix)
        username = event_data.get('username', '')
        if username and not username.startswith('@'):
            username = f'@{username}'
        
        # Format inviter name (prefer username over name) - handle Unknown case
        inviter_name = event_data.get('inviter_name', '')
        if inviter_name == 'Unknown' or not inviter_name:
            inviter_name = 'Не указан'
        
        # Add data to row according to TZ format
        row_data = [
            formatted_time,
            action,
            event_data.get('tg_user_id', ''),
            username,
            event_data.get('user_name', event_data.get('name', '')),
            inviter_name,
            status
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        wb.save(self.excel_file)
        logger.info(f"Added history event: {action} for user {event_data.get('tg_user_id')}")
    
    def update_statistics_sheet(self) -> None:
        """Update Статистика sheet with current data from database."""
        wb = load_workbook(self.excel_file)
        ws = wb["Статистика"]
        
        # Clear existing data (keep headers)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # Get current statistics from database
        stats_data = self._get_current_statistics()
        
        # Add data to sheet with proper name handling
        for row_idx, inviter_stats in enumerate(stats_data, 2):
            # Normalize inviter name - translate Unknown to Russian
            inviter_name = inviter_stats.get('inviter_name', 'Не указан')
            if inviter_name == 'Unknown' or not inviter_name:
                inviter_name = 'Не указан'
            
            ws.cell(row=row_idx, column=1, value=inviter_name)
            ws.cell(row=row_idx, column=2, value=inviter_stats.get('total_invited', 0))
            ws.cell(row=row_idx, column=3, value=inviter_stats.get('currently_subscribed', 0))
            ws.cell(row=row_idx, column=4, value=inviter_stats.get('unsubscribed', 0))
            
            # Calculate retention percentage with safe access
            total_invited = inviter_stats.get('total_invited', 0)
            currently_subscribed = inviter_stats.get('currently_subscribed', 0)
            retention_pct = 0
            if total_invited > 0:
                retention_pct = round((currently_subscribed / total_invited) * 100)
            ws.cell(row=row_idx, column=5, value=f"{retention_pct}%")
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        wb.save(self.excel_file)
        logger.info("Updated Статистика sheet")
    
    def _get_current_statistics(self) -> List[Dict[str, Any]]:
        """Get current statistics by inviter from database - use proper username priority."""
        try:
            # Use database method that properly handles username > name priority
            stats_data = self.db.get_statistics_data()
            
            # Normalize all inviter names to handle Unknown values
            for stat in stats_data:
                inviter_name = stat.get('inviter_name', 'Не указан')
                if inviter_name == 'Unknown' or not inviter_name:
                    stat['inviter_name'] = 'Не указан'
            
            return stats_data
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return []
    
    def create_daily_report_sheet(self, target_date: date) -> Dict[str, Any]:
        """Create daily report sheet with format ДД-ММ-ГГГГ."""
        sheet_name = target_date.strftime("%d-%m-%Y")
        
        wb = load_workbook(self.excel_file)
        
        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            logger.info(f"Daily sheet {sheet_name} already exists")
            return {'sheet_exists': True, 'sheet_name': sheet_name}
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Get daily statistics
        daily_stats = self._get_daily_statistics(target_date)
        
        # Create the report format according to TZ
        current_row = 1
        
        # Title
        title_cell = ws.cell(row=current_row, column=1, value="Ежедневный отчёт")
        title_cell.font = Font(size=14, bold=True)
        current_row += 2
        
        # Data rows according to TZ format
        report_data = [
            ("Дата", target_date.strftime("%d.%m.%Y")),
            ("Новые пользователи", daily_stats['new_users']),
            ("Остались на " + (target_date + timedelta(days=3)).strftime("%d.%m"), 
             f"{daily_stats['retained_3_days']} ({daily_stats['retention_3_days_pct']}%)"),
            ("Вышли в тот же день", daily_stats['left_same_day']),
            ("Топ-3 пригласителей", daily_stats['top_inviters_text']),
            ("Динамика (vs вчера)", daily_stats['dynamics_text'])
        ]
        
        for param, value in report_data:
            ws.cell(row=current_row, column=1, value=param).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        wb.save(self.excel_file)
        logger.info(f"Created daily report sheet: {sheet_name}")
        
        return {
            'sheet_exists': False, 
            'sheet_name': sheet_name,
            'stats': daily_stats,
            'excel_file': str(self.excel_file)
        }
    
    def _get_daily_statistics(self, target_date: date) -> Dict[str, Any]:
        """Get daily statistics for the specified date."""
        try:
            # Get basic stats for the day
            date_str = target_date.strftime("%Y-%m-%d")
            daily_stats = self.db.get_daily_stats(date_str)
            
            new_users = daily_stats.get('total_subscriptions', 0)
            left_same_day = daily_stats.get('same_day_unsubscriptions', 0)
            
            # Calculate 3-day retention
            retention_3_days = self.db.get_retention_for_date(date_str, 3)
            retained_3_days = retention_3_days.get('retained', 0)
            retention_3_days_pct = round(retention_3_days.get('retention_rate', 0))
            
            # Get top 3 inviters for the day
            top_inviters = self._get_top_inviters_for_date(target_date)
            top_inviters_text = self._format_top_inviters(top_inviters)
            
            # Get dynamics (compare with yesterday)
            yesterday = target_date - timedelta(days=1)
            yesterday_stats = self.db.get_daily_stats(yesterday.strftime("%Y-%m-%d"))
            dynamics_text = self._format_dynamics(daily_stats, yesterday_stats)
            
            return {
                'new_users': new_users,
                'retained_3_days': retained_3_days,
                'retention_3_days_pct': retention_3_days_pct,
                'left_same_day': left_same_day,
                'top_inviters_text': top_inviters_text,
                'dynamics_text': dynamics_text,
                'top_inviters': top_inviters
            }
            
        except Exception as e:
            logger.error(f"Error getting daily statistics: {e}")
            return {
                'new_users': 0,
                'retained_3_days': 0,
                'retention_3_days_pct': 0,
                'left_same_day': 0,
                'top_inviters_text': 'Нет данных',
                'dynamics_text': 'Нет данных',
                'top_inviters': []
            }
    
    def _get_top_inviters_for_date(self, target_date: date) -> List[Dict[str, Any]]:
        """Get top 3 inviters for specific date."""
        try:
            date_str = target_date.strftime("%Y-%m-%d")
            return self.db.get_top_inviters_for_date(date_str, limit=3)
        except Exception as e:
            logger.error(f"Error getting top inviters: {e}")
            return []
    
    def _format_top_inviters(self, top_inviters: List[Dict[str, Any]]) -> str:
        """Format top inviters according to TZ specification."""
        if not top_inviters:
            return "Нет данных"
        
        formatted_lines = []
        for i, inviter in enumerate(top_inviters, 1):
            # Normalize inviter name - translate Unknown to Russian
            name = inviter.get('inviter_name', 'Не указан')
            if name == 'Unknown' or not name:
                name = 'Не указан'
            invited = inviter.get('invited_count', 0)
            retained = inviter.get('retained_count', 0)
            retention_pct = 0
            if invited > 0:
                retention_pct = round((retained / invited) * 100)
            
            formatted_lines.append(f"{name} — {invited} ({retained} удержано, {retention_pct}%)")
        
        return ", ".join(formatted_lines)
    
    def _format_dynamics(self, today_stats: Dict[str, Any], yesterday_stats: Dict[str, Any]) -> str:
        """Format dynamics comparison according to TZ specification."""
        today_new = today_stats.get('total_subscriptions', 0)
        today_retention = round(today_stats.get('retention_rate', 0))
        
        yesterday_new = yesterday_stats.get('total_subscriptions', 0)
        yesterday_retention = round(yesterday_stats.get('retention_rate', 0))
        
        return f"Вчера: +{yesterday_new} новых, {yesterday_retention}% удержание, Сегодня: +{today_new} новых, {today_retention}% удержание"
    
    def get_daily_report_message(self, target_date: date) -> str:
        """Generate daily report message according to TZ specification."""
        daily_stats = self._get_daily_statistics(target_date)
        
        # Format date according to TZ
        formatted_date = target_date.strftime("%d.%m.%Y")
        
        # Top 3 inviters formatted for message
        top_inviters_lines = []
        for i, inviter in enumerate(daily_stats['top_inviters'][:3], 1):
            name = inviter.get('inviter_name', 'Unknown')
            invited = inviter.get('invited_count', 0)
            retained = inviter.get('retained_count', 0)
            retention_pct = 0
            if invited > 0:
                retention_pct = round((retained / invited) * 100)
            
            top_inviters_lines.append(f"{i}. {name} — {invited} ({retained} удержано, {retention_pct}%)")
        
        top_inviters_text = "\n".join(top_inviters_lines) if top_inviters_lines else "Нет данных"
        
        message = f"""📊 Ежедневный отчёт — {formatted_date}

👥 Новые пользователи: {daily_stats['new_users']}
✅ Остались на {(target_date + timedelta(days=3)).strftime('%d.%m')}: {daily_stats['retained_3_days']} ({daily_stats['retention_3_days_pct']}%)
❌ Вышли в тот же день: {daily_stats['left_same_day']}

🏆 Топ-3 пригласителей:
{top_inviters_text}

📈 Динамика:
{daily_stats['dynamics_text']}

📤 Скачать Excel-файл"""

        return message
    
    def get_excel_file_path(self) -> str:
        """Get path to the unified Excel file."""
        return str(self.excel_file)