#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Templates module for Telegram bot reporting system.
Provides templates and utilities for generating formatted Excel reports
with proper styling, charts, and Russian localization.
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
# from openpyxl.chart.series import DataPoint  # Not needed

from utils.time_utils import format_datetime_for_report, format_time_period_ru

logger = logging.getLogger(__name__)


class BaseReportTemplate:
    """
    Base class for Excel report templates with common styling and utilities.
    """
    
    def __init__(self):
        """Initialize base template with common styles."""
        self.wb = None
        self._init_styles()
    
    def _init_styles(self):
        """Initialize common Excel styles."""
        # Header style
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, size=12, color="FFFFFF")
        self.header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Title style
        self.title_style = NamedStyle(name="title")
        self.title_style.font = Font(bold=True, size=16, color="366092")
        self.title_style.alignment = Alignment(horizontal="center")
        
        # Subtitle style
        self.subtitle_style = NamedStyle(name="subtitle")
        self.subtitle_style.font = Font(bold=True, size=12, color="666666")
        self.subtitle_style.alignment = Alignment(horizontal="left")
        
        # Data style
        self.data_style = NamedStyle(name="data")
        self.data_style.font = Font(size=10)
        self.data_style.alignment = Alignment(horizontal="center")
        self.data_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Number style
        self.number_style = NamedStyle(name="number")
        self.number_style.font = Font(size=10)
        self.number_style.alignment = Alignment(horizontal="right")
        self.number_style.number_format = "#,##0"
        
        # Percentage style
        self.percentage_style = NamedStyle(name="percentage")
        self.percentage_style.font = Font(size=10)
        self.percentage_style.alignment = Alignment(horizontal="right")
        self.percentage_style.number_format = "0.00%"
    
    def _add_title(self, ws, title: str, row: int = 1) -> int:
        """
        Add title to worksheet.
        
        Args:
            ws: Worksheet
            title: Title text
            row: Starting row
            
        Returns:
            int: Next available row
        """
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).style = self.title_style
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 2
    
    def _add_subtitle(self, ws, subtitle: str, row: int) -> int:
        """
        Add subtitle to worksheet.
        
        Args:
            ws: Worksheet
            subtitle: Subtitle text
            row: Starting row
            
        Returns:
            int: Next available row
        """
        ws.cell(row=row, column=1, value=subtitle)
        ws.cell(row=row, column=1).style = self.subtitle_style
        return row + 1
    
    def _add_dataframe_table(self, ws, df: pd.DataFrame, start_row: int, 
                            start_col: int = 1, add_headers: bool = True) -> int:
        """
        Add DataFrame as formatted table to worksheet.
        
        Args:
            ws: Worksheet
            df: DataFrame to add
            start_row: Starting row
            start_col: Starting column
            add_headers: Whether to add headers
            
        Returns:
            int: Next available row
        """
        current_row = start_row
        
        # Add headers if requested
        if add_headers:
            for col_idx, column in enumerate(df.columns):
                cell = ws.cell(row=current_row, column=start_col + col_idx, value=column)
                cell.style = self.header_style
            current_row += 1
        
        # Add data rows
        for _, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=current_row, column=start_col + col_idx, value=value)
                cell.style = self.data_style
                
                # Apply number formatting for numeric values
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    if 0 <= value <= 1 and '.' in str(value):
                        cell.style = self.percentage_style
                    else:
                        cell.style = self.number_style
            current_row += 1
        
        return current_row + 1
    
    def _add_summary_stats(self, ws, stats: Dict[str, Any], start_row: int) -> int:
        """
        Add summary statistics section.
        
        Args:
            ws: Worksheet
            stats: Statistics dictionary
            start_row: Starting row
            
        Returns:
            int: Next available row
        """
        current_row = start_row
        
        # Define stat mappings with Russian labels
        stat_mappings = {
            'total_subscriptions': 'Всего подписок',
            'total_unsubscriptions': 'Всего отписок',
            'net_growth': 'Чистый прирост',
            'unique_subscribers': 'Уникальных подписчиков',
            'repeat_subscribers': 'Повторных подписчиков',
            'retention_rate': 'Показатель удержания',
            'churn_rate': 'Показатель оттока'
        }
        
        for key, label in stat_mappings.items():
            if key in stats:
                value = stats[key]
                
                # Format value appropriately
                if isinstance(value, float) and 0 <= value <= 1:
                    formatted_value = f"{value:.2%}"
                elif isinstance(value, (int, float)):
                    formatted_value = f"{value:,}"
                else:
                    formatted_value = str(value)
                
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=1).style = self.subtitle_style
                ws.cell(row=current_row, column=2, value=formatted_value)
                ws.cell(row=current_row, column=2).style = self.data_style
                
                current_row += 1
        
        return current_row + 1
    
    def _add_bar_chart(self, ws, data_range: str, categories_range: str, 
                       title: str, position: str) -> None:
        """
        Add bar chart to worksheet.
        
        Args:
            ws: Worksheet
            data_range: Data range for chart
            categories_range: Categories range
            title: Chart title
            position: Chart position (e.g., "H5")
        """
        try:
            chart = BarChart()
            chart.title = title
            chart.y_axis.title = "Количество"
            chart.x_axis.title = "Период"
            
            # Create Reference objects properly
            data = Reference(worksheet=ws, range_string=data_range)
            categories = Reference(worksheet=ws, range_string=categories_range)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, position)
        except Exception as e:
            logger.warning(f"Failed to create bar chart: {e}")
            # Chart creation failed, continue without chart
    
    def _add_line_chart(self, ws, data_range: str, categories_range: str,
                        title: str, position: str) -> None:
        """
        Add line chart to worksheet.
        
        Args:
            ws: Worksheet
            data_range: Data range for chart
            categories_range: Categories range
            title: Chart title
            position: Chart position (e.g., "H5")
        """
        try:
            chart = LineChart()
            chart.title = title
            chart.y_axis.title = "Значение"
            chart.x_axis.title = "Дата"
            
            # Create Reference objects properly
            data = Reference(worksheet=ws, range_string=data_range)
            categories = Reference(worksheet=ws, range_string=categories_range)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, position)
        except Exception as e:
            logger.warning(f"Failed to create line chart: {e}")
            # Chart creation failed, continue without chart
    
    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width


class DailyReportTemplate(BaseReportTemplate):
    """Template for daily subscription/unsubscription reports."""
    
    def generate(self, report_data: Dict[str, Any], file_path: Path) -> None:
        """
        Generate daily report Excel file.
        
        Args:
            report_data: Report data dictionary
            file_path: Output file path
        """
        logger.info(f"Generating daily report: {file_path}")
        
        self.wb = Workbook()
        # Remove default sheet safely
        active = self.wb.active
        if active is not None:
            self.wb.remove(active)
        
        # Add styles to workbook
        for style in [self.header_style, self.title_style, self.subtitle_style, 
                     self.data_style, self.number_style, self.percentage_style]:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
        
        # Create main summary sheet
        self._create_summary_sheet(report_data)
        
        # Create events detail sheet
        self._create_events_sheet(report_data)
        
        # Create retention analysis sheet
        self._create_retention_sheet(report_data)
        
        # Save workbook
        self.wb.save(file_path)
        logger.info(f"Daily report saved: {file_path}")
    
    def _create_summary_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create summary sheet with main statistics."""
        ws = self.wb.create_sheet("Сводка")
        
        # Title
        target_date = report_data['date']
        formatted_date = format_datetime_for_report(
            datetime.fromisoformat(target_date + "T00:00:00"), 
            include_time=False
        )
        current_row = self._add_title(ws, f"Ежедневный отчёт за {formatted_date}")
        
        # Summary statistics
        current_row = self._add_subtitle(ws, "Основная статистика", current_row)
        current_row = self._add_summary_stats(ws, report_data['stats'], current_row)
        
        # Retention statistics
        current_row = self._add_subtitle(ws, "Показатели удержания", current_row + 1)
        
        retention_data = []
        for period, retention_stats in report_data['retention'].items():
            period_days = period.replace('_', ' ')
            retention_data.append({
                'Период': format_time_period_ru(int(period.split('_')[0])),
                'Всего подписок': retention_stats.get('total_subscriptions', 0),
                'Удержано': retention_stats.get('retained', 0),
                'Не удержано': retention_stats.get('not_retained', 0),
                'Процент удержания': retention_stats.get('retention_rate', 0)
            })
        
        if retention_data:
            df_retention = pd.DataFrame(retention_data)
            current_row = self._add_dataframe_table(ws, df_retention, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_events_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create events detail sheet."""
        ws = self.wb.create_sheet("События")
        
        # Title
        current_row = self._add_title(ws, "Детализация событий")
        
        # Events table
        events = report_data.get('events', [])
        if events:
            # Convert events to DataFrame
            events_df = pd.DataFrame(events)
            
            # Rename columns to Russian
            column_mapping = {
                'event_time': 'Время события',
                'event_type': 'Тип события',
                'tg_user_id': 'ID пользователя',
                'username': 'Имя пользователя',
                'name': 'Отображаемое имя',
                'status': 'Статус',
                'note': 'Примечание'
            }
            
            events_df = events_df.rename(columns=column_mapping)
            
            # Format datetime columns
            if 'Время события' in events_df.columns:
                events_df['Время события'] = events_df['Время события'].apply(
                    lambda x: format_datetime_for_report(datetime.fromisoformat(x)) if x else ''
                )
            
            current_row = self._add_dataframe_table(ws, events_df, current_row)
        else:
            ws.cell(row=current_row, column=1, value="Нет событий за данный период")
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_retention_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create retention analysis sheet."""
        ws = self.wb.create_sheet("Анализ удержания")
        
        # Title
        current_row = self._add_title(ws, "Анализ удержания пользователей")
        
        # Detailed retention analysis
        for period, retention_stats in report_data['retention'].items():
            if not retention_stats:
                continue
            
            period_name = format_time_period_ru(int(period.split('_')[0]))
            current_row = self._add_subtitle(ws, f"Удержание за {period_name}", current_row)
            
            # Add retention statistics
            retention_summary = {
                'total_subscriptions': retention_stats.get('total_subscriptions', 0),
                'retained': retention_stats.get('retained', 0),
                'not_retained': retention_stats.get('not_retained', 0),
                'retention_rate': retention_stats.get('retention_rate', 0)
            }
            
            current_row = self._add_summary_stats(ws, retention_summary, current_row)
            current_row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)


class WeeklyReportTemplate(BaseReportTemplate):
    """Template for weekly subscription/unsubscription reports."""
    
    def generate(self, report_data: Dict[str, Any], file_path: Path) -> None:
        """Generate weekly report Excel file."""
        logger.info(f"Generating weekly report: {file_path}")
        
        self.wb = Workbook()
        # Remove default sheet safely
        active = self.wb.active
        if active is not None:
            self.wb.remove(active)
        
        # Add styles
        for style in [self.header_style, self.title_style, self.subtitle_style, 
                     self.data_style, self.number_style, self.percentage_style]:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
        
        # Create sheets
        self._create_weekly_summary_sheet(report_data)
        self._create_daily_breakdown_sheet(report_data)
        self._create_weekly_events_sheet(report_data)
        
        # Save workbook
        self.wb.save(file_path)
        logger.info(f"Weekly report saved: {file_path}")
    
    def _create_weekly_summary_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create weekly summary sheet."""
        ws = self.wb.create_sheet("Недельная сводка")
        
        # Title
        week_start = report_data['week_start']
        week_end = report_data['week_end']
        current_row = self._add_title(ws, f"Недельный отчёт с {week_start} по {week_end}")
        
        # Weekly statistics
        current_row = self._add_subtitle(ws, "Статистика за неделю", current_row)
        current_row = self._add_summary_stats(ws, report_data['stats'], current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_daily_breakdown_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create daily breakdown sheet."""
        ws = self.wb.create_sheet("По дням")
        
        # Title
        current_row = self._add_title(ws, "Разбивка по дням недели")
        
        # Daily breakdown table
        daily_data = []
        for day_data in report_data.get('daily_breakdown', []):
            daily_data.append({
                'Дата': day_data['date'],
                'День недели': day_data['weekday'],
                'Подписки': day_data['stats'].get('total_subscriptions', 0),
                'Отписки': day_data['stats'].get('total_unsubscriptions', 0),
                'Чистый прирост': day_data['stats'].get('net_growth', 0)
            })
        
        if daily_data:
            df_daily = pd.DataFrame(daily_data)
            current_row = self._add_dataframe_table(ws, df_daily, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_weekly_events_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create weekly events sheet."""
        ws = self.wb.create_sheet("События недели")
        
        # Title
        current_row = self._add_title(ws, "Все события за неделю")
        
        # Events table (similar to daily report)
        events = report_data.get('events', [])
        if events:
            events_df = pd.DataFrame(events)
            
            column_mapping = {
                'event_time': 'Время события',
                'event_type': 'Тип события',
                'tg_user_id': 'ID пользователя',
                'username': 'Имя пользователя',
                'status': 'Статус'
            }
            
            events_df = events_df.rename(columns=column_mapping)
            
            if 'Время события' in events_df.columns:
                events_df['Время события'] = events_df['Время события'].apply(
                    lambda x: format_datetime_for_report(datetime.fromisoformat(x)) if x else ''
                )
            
            current_row = self._add_dataframe_table(ws, events_df, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)


class MonthlyReportTemplate(BaseReportTemplate):
    """Template for monthly subscription/unsubscription reports."""
    
    def generate(self, report_data: Dict[str, Any], file_path: Path) -> None:
        """Generate monthly report Excel file."""
        logger.info(f"Generating monthly report: {file_path}")
        
        self.wb = Workbook()
        # Remove default sheet safely
        active = self.wb.active
        if active is not None:
            self.wb.remove(active)
        
        # Add styles
        for style in [self.header_style, self.title_style, self.subtitle_style, 
                     self.data_style, self.number_style, self.percentage_style]:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
        
        # Create sheets
        self._create_monthly_summary_sheet(report_data)
        self._create_weekly_breakdown_sheet(report_data)
        self._create_monthly_retention_sheet(report_data)
        
        # Save workbook
        self.wb.save(file_path)
        logger.info(f"Monthly report saved: {file_path}")
    
    def _create_monthly_summary_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create monthly summary sheet."""
        ws = self.wb.create_sheet("Месячная сводка")
        
        # Title
        month_start = report_data['month_start']
        month_end = report_data['month_end']
        current_row = self._add_title(ws, f"Месячный отчёт с {month_start} по {month_end}")
        
        # Monthly statistics
        current_row = self._add_subtitle(ws, "Статистика за месяц", current_row)
        current_row = self._add_summary_stats(ws, report_data['stats'], current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_weekly_breakdown_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create weekly breakdown sheet."""
        ws = self.wb.create_sheet("По неделям")
        
        # Title
        current_row = self._add_title(ws, "Разбивка по неделям")
        
        # Weekly breakdown table
        weekly_data = []
        for week_data in report_data.get('weekly_breakdown', []):
            weekly_data.append({
                'Начало недели': week_data['week_start'],
                'Конец недели': week_data['week_end'],
                'Подписки': week_data['stats'].get('total_subscriptions', 0),
                'Отписки': week_data['stats'].get('total_unsubscriptions', 0),
                'Чистый прирост': week_data['stats'].get('net_growth', 0)
            })
        
        if weekly_data:
            df_weekly = pd.DataFrame(weekly_data)
            current_row = self._add_dataframe_table(ws, df_weekly, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_monthly_retention_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create monthly retention analysis sheet."""
        ws = self.wb.create_sheet("Удержание")
        
        # Title
        current_row = self._add_title(ws, "Анализ удержания за месяц")
        
        # Retention analysis
        retention_analysis = report_data.get('retention_analysis', {})
        for period, retention_stats in retention_analysis.items():
            if not retention_stats:
                continue
            
            period_name = format_time_period_ru(int(period.split('_')[0]))
            current_row = self._add_subtitle(ws, f"Удержание за {period_name}", current_row)
            current_row = self._add_summary_stats(ws, retention_stats, current_row)
            current_row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)


class RetentionReportTemplate(BaseReportTemplate):
    """Template for detailed retention analysis reports."""
    
    def generate(self, report_data: Dict[str, Any], file_path: Path) -> None:
        """Generate retention report Excel file."""
        logger.info(f"Generating retention report: {file_path}")
        
        self.wb = Workbook()
        # Remove default sheet safely
        active = self.wb.active
        if active is not None:
            self.wb.remove(active)
        
        # Add styles
        for style in [self.header_style, self.title_style, self.subtitle_style, 
                     self.data_style, self.number_style, self.percentage_style]:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
        
        # Create sheets
        self._create_retention_summary_sheet(report_data)
        self._create_retention_details_sheet(report_data)
        self._create_retention_trends_sheet(report_data)
        
        # Save workbook
        self.wb.save(file_path)
        logger.info(f"Retention report saved: {file_path}")
    
    def _create_retention_summary_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create retention summary sheet."""
        ws = self.wb.create_sheet("Сводка удержания")
        
        # Title
        retention_days = report_data['retention_days']
        target_date = report_data['target_date']
        period_name = format_time_period_ru(retention_days)
        
        current_row = self._add_title(ws, f"Анализ удержания за {period_name} на {target_date}")
        
        # Summary statistics
        current_row = self._add_subtitle(ws, "Общая статистика", current_row)
        current_row = self._add_summary_stats(ws, report_data['stats'], current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_retention_details_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create retention details sheet."""
        ws = self.wb.create_sheet("Детали удержания")
        
        # Title
        current_row = self._add_title(ws, "Детальный анализ удержания")
        
        # Details table
        details = report_data.get('details', [])
        if details:
            details_df = pd.DataFrame(details)
            
            column_mapping = {
                'journal_id': 'ID записи',
                'tg_user_id': 'ID пользователя',
                'username': 'Имя пользователя',
                'subscription_date': 'Дата подписки',
                'retention_result': 'Результат удержания',
                'inviter_id': 'ID пригласившего'
            }
            
            details_df = details_df.rename(columns=column_mapping)
            
            if 'Дата подписки' in details_df.columns:
                details_df['Дата подписки'] = details_df['Дата подписки'].apply(
                    lambda x: format_datetime_for_report(datetime.fromisoformat(x)) if x else ''
                )
            
            current_row = self._add_dataframe_table(ws, details_df, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_retention_trends_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create retention trends sheet."""
        ws = self.wb.create_sheet("Тренды удержания")
        
        # Title
        current_row = self._add_title(ws, "Тренды удержания по дням")
        
        # Trends table
        trends = report_data.get('trends', [])
        if trends:
            trends_data = []
            for trend in trends:
                trends_data.append({
                    'Дата': trend['date'],
                    'Всего подписок': trend['stats'].get('total_subscriptions', 0),
                    'Удержано': trend['stats'].get('retained', 0),
                    'Процент удержания': trend['stats'].get('retention_rate', 0)
                })
            
            if trends_data:
                df_trends = pd.DataFrame(trends_data)
                current_row = self._add_dataframe_table(ws, df_trends, current_row)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)